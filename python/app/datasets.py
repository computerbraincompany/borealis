"""Tabular dataset registry backed by DuckDB.

Every uploaded/connected tabular file becomes a DuckDB table owned by an account.
The account namespace prevents cross-user SQL access: registrations are keyed by
(account_id, name) and the in-memory catalog is only reachable through /query
with a matching account_id. DuckDB runs embedded in-memory; datasets are
re-registered from disk on startup.
"""

from __future__ import annotations

import csv
import json
import logging
import math
import re
import tempfile
import threading
import zipfile
from collections import OrderedDict
from collections.abc import Sequence
from decimal import Decimal
from pathlib import Path
from typing import Any

import duckdb
from fastapi import HTTPException
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
LOCK = threading.RLock()
# Dataset work is coordinated per account so a slow query or catalog load for
# one tenant cannot stall every other tenant in the process.  ``LOCK`` remains
# a deliberately narrow lock for the shared lock/cache maps (and as a stable
# test/debugging hook); it is never held while DuckDB or network I/O runs.
_ACCOUNT_LOCKS: dict[str, threading.RLock] = {}

# account_id -> { table_name: {meta} }
_REGISTRY: dict[str, dict[str, dict[str, Any]]] = {}
_PENDING_ACTIVATIONS: dict[tuple[str, str, str], int] = {}
_PENDING_PREPARATIONS: dict[tuple[str, str, str], int] = {}
# (account_id, sorted allowed tables) -> isolated in-memory DuckDB catalog
ScopeKey = tuple[str, tuple[str, ...]]
ScopeSignatures = tuple[tuple[str, str], ...]
_CONNECTIONS: OrderedDict[ScopeKey, duckdb.DuckDBPyConnection] = OrderedDict()
_SCOPE_SIGNATURES: dict[ScopeKey, ScopeSignatures] = {}
MAX_SCOPES_PER_ACCOUNT = 8
MAX_DUCKDB_QUERY_SECONDS = 30.0
DUCKDB_THREADS = 4
DUCKDB_MEMORY_LIMIT = "512MB"
DUCKDB_TEMP_LIMIT = "512MB"
MAX_QUERY_ROWS = 500
MAX_QUERY_COLUMNS = 100
MAX_QUERY_CELLS = 50_000
MAX_QUERY_CHARS = 1_000_000
MAX_QUERY_CELL_CHARS = 10_000
MAX_QUERY_HEADER_CHARS = 100_000
MAX_QUERY_COLUMN_NAME_CHARS = 500
MAX_EXTRACT_ROWS = 2_000
MAX_EXTRACT_CELLS = 50_000
MAX_EXTRACT_CHARS = 1_000_000
MAX_EXTRACT_CELL_CHARS = 10_000
MAX_EXTRACT_COLUMNS = 500
MAX_EXTRACT_HEADER_CHARS = 100_000
MAX_EXTRACT_COLUMN_NAME_CHARS = 500
MAX_DESCRIBE_ROWS = 100_000
MAX_DESCRIBE_COLUMNS = 100
MAX_TOP_VALUE_COLUMNS = 20
MAX_DESCRIBE_VALUE_CHARS = 500
MAX_DESCRIBE_CHARS = 128_000
MAX_DATASET_COLUMNS = 500
MAX_DATASET_COLUMN_NAME_CHARS = 500
MAX_DATASET_TYPE_CHARS = 500
MAX_PREVIEW_ROWS = 5
MAX_PREVIEW_CELL_CHARS = 500
MAX_PREVIEW_CHARS = 100_000
MAX_CATALOG_CHARS = 256_000
MAX_XLSX_ROWS = 200_000
MAX_XLSX_COLUMNS = 10_000
MAX_XLSX_CELLS = 2_000_000
MAX_XLSX_EXPANDED_BYTES = 100 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 50 * 1024 * 1024
MAX_XLSX_ARCHIVE_MEMBERS = 10_000
MAX_XLSX_CELL_CHARS = 1_000_000
JS_MAX_SAFE_INTEGER = 9_007_199_254_740_991

TABLE_RE = re.compile(r"^[a-z][a-z0-9_]{0,62}$")
EUROPEAN_DATE_RE = re.compile(r"(?<!\d)\d{2}\.\d{2}\.\d{2,4}(?!\d)")
LOGGER = logging.getLogger(__name__)


def _account_lock(account_id: str) -> threading.RLock:
    with LOCK:
        return _ACCOUNT_LOCKS.setdefault(account_id, threading.RLock())


def _is_european_semicolon_csv(sample: str) -> bool:
    """Conservatively detect semicolon CSVs containing European-style dates."""
    delimited_lines = [line for line in sample.lstrip("\ufeff").splitlines() if line.count(";") >= 2]
    if len(delimited_lines) < 2:
        return False
    delimiter_counts = [line.count(";") for line in delimited_lines]
    has_consistent_rows = any(delimiter_counts.count(count) >= 2 for count in set(delimiter_counts))
    return has_consistent_rows and any(EUROPEAN_DATE_RE.search(line) for line in delimited_lines)


def _read_sql(path: str, expected_format: str | None = None) -> str:
    ext = Path(path).suffix.lower()
    if ext == ".xls":
        raise HTTPException(422, "legacy .xls spreadsheets are not supported; use .xlsx")
    if ext == ".xlsx":
        raise HTTPException(422, "xlsx input must be converted before loading")
    if expected_format == "json":
        return "read_json_auto(?)"
    if expected_format == "csv":
        with Path(path).open("rb") as source:
            sample = source.read(4096).decode("utf-8-sig", errors="ignore")
        if _is_european_semicolon_csv(sample):
            return (
                "read_csv_auto(?, delim=';', dateformat='%d.%m.%y', decimal_separator=',', "
                "thousands='.', auto_type_candidates=['BOOLEAN','BIGINT','DECIMAL','DATE','TIMESTAMP','VARCHAR'])"
            )
        return "read_csv_auto(?)"
    if ext in (".parquet",):
        return "read_parquet(?)"
    if ext in (".json", ".jsonl"):
        return "read_json_auto(?)"
    if ext in (".tsv",):
        return "read_csv_auto(?, delim='\\t')"
    if ext == ".csv":
        # Do not read a potentially large connector into memory just to sniff
        # its delimiter/date convention.
        with Path(path).open("rb") as source:
            sample = source.read(4096).decode("utf-8-sig", errors="ignore")
        if _is_european_semicolon_csv(sample):
            return (
                "read_csv_auto(?, delim=';', dateformat='%d.%m.%y', decimal_separator=',', "
                "thousands='.', auto_type_candidates=['BOOLEAN','BIGINT','DECIMAL','DATE','TIMESTAMP','VARCHAR'])"
            )
    return "read_csv_auto(?)"


def _validate_expected_file_format(path: str, expected_format: str | None) -> None:
    if expected_format not in {"csv", "json"}:
        return
    with Path(path).open("rb") as source:
        prefix = source.read(512).lstrip(b"\xef\xbb\xbf \t\r\n")
    if not prefix:
        raise HTTPException(422, "dataset file is empty")
    lowered = prefix.lower()
    if lowered.startswith((b"<!doctype html", b"<html")):
        raise HTTPException(422, "dataset file contains HTML, not tabular data")
    looks_json = prefix.startswith((b"{", b"["))
    if expected_format == "json" and not looks_json:
        raise HTTPException(422, "dataset does not match expected JSON format")
    if expected_format == "csv" and looks_json:
        raise HTTPException(422, "dataset does not match expected CSV format")


def _xlsx_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bytes):
        return value.hex()
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (TypeError, ValueError):
            pass
    return value


def _xlsx_to_csv(path: str) -> Path:
    """Convert the first worksheet to a bounded temporary CSV without network extensions."""
    workbook = None
    output_path: Path | None = None
    completed = False
    try:
        try:
            with zipfile.ZipFile(path) as archive:
                members = archive.infolist()
                if len(members) > MAX_XLSX_ARCHIVE_MEMBERS:
                    raise HTTPException(413, "xlsx archive has too many members")
                expanded_bytes = 0
                for member in members:
                    if member.flag_bits & 0x1:
                        raise HTTPException(422, "encrypted xlsx workbooks are not supported")
                    if member.file_size > MAX_XLSX_MEMBER_BYTES:
                        raise HTTPException(413, "xlsx archive member exceeds the processing limit")
                    expanded_bytes += member.file_size
                    if expanded_bytes > MAX_XLSX_EXPANDED_BYTES:
                        raise HTTPException(413, "xlsx archive expands beyond the processing limit")
        except zipfile.BadZipFile as exc:
            raise HTTPException(422, "xlsx workbook could not be parsed") from exc

        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        if not workbook.worksheets:
            raise HTTPException(422, "xlsx workbook has no worksheets")
        worksheet = workbook.worksheets[0]
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            newline="",
            prefix="borealis-xlsx-",
            suffix=".csv",
            delete=False,
        ) as output:
            output_path = Path(output.name)
            writer = csv.writer(output)
            row_count = 0
            cell_count = 0
            for values in worksheet.iter_rows(values_only=True):
                row = list(values)
                while row and row[-1] is None:
                    row.pop()
                if len(row) > MAX_XLSX_COLUMNS:
                    raise HTTPException(413, "xlsx worksheet has too many columns")
                row_count += 1
                cell_count += len(row)
                if row_count > MAX_XLSX_ROWS or cell_count > MAX_XLSX_CELLS:
                    raise HTTPException(413, "xlsx worksheet exceeds the processing limit")
                rendered_row = [_xlsx_cell(value) for value in row]
                encoded_lengths = [len(str(value).encode("utf-8")) for value in rendered_row]
                if any(length > MAX_XLSX_CELL_CHARS for length in encoded_lengths):
                    raise HTTPException(413, "xlsx cell exceeds the processing limit")
                # CSV quoting can at most double every byte (all quotes), plus
                # a small delimiter/quote/newline overhead per field.
                row_upper_bound = sum(length * 2 + 3 for length in encoded_lengths) + 2
                if output.tell() + row_upper_bound > MAX_XLSX_EXPANDED_BYTES:
                    raise HTTPException(413, "xlsx worksheet expands beyond the processing limit")
                writer.writerow(rendered_row)
                if output.tell() > MAX_XLSX_EXPANDED_BYTES:
                    raise HTTPException(413, "xlsx worksheet expands beyond the processing limit")
        if row_count == 0:
            raise HTTPException(422, "xlsx worksheet is empty")
        completed = True
        return output_path
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(422, "xlsx workbook could not be parsed") from exc
    finally:
        if workbook is not None:
            workbook.close()
        if output_path is not None and not completed:
            output_path.unlink(missing_ok=True)


def _load_table(
    con: duckdb.DuckDBPyConnection,
    name: str,
    location: str,
    expected_format: str | None = None,
) -> None:
    source_path = location
    temporary_path: Path | None = None
    try:
        if Path(location).suffix.lower() == ".xlsx":
            temporary_path = _xlsx_to_csv(location)
            source_path = str(temporary_path)
        con.execute(
            f"CREATE TABLE {_qi(name)} AS SELECT * FROM {_read_sql(source_path, expected_format)}",
            [source_path],
        )
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def _qi(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def _configure_connection(con: duckdb.DuckDBPyConnection) -> None:
    """Apply trusted per-catalog resource limits before any uploaded SQL runs."""
    con.execute(f"SET threads={DUCKDB_THREADS}")
    con.execute(f"SET memory_limit='{DUCKDB_MEMORY_LIMIT}'")
    con.execute(f"SET max_temp_directory_size='{DUCKDB_TEMP_LIMIT}'")


def _file_sig(path: str) -> str:
    try:
        stat = Path(path).stat()
        return f"{stat.st_size}:{stat.st_mtime_ns}"
    except OSError:
        return "missing"


def _scope_key(account_id: str, allowed_tables: Sequence[str]) -> ScopeKey:
    if isinstance(allowed_tables, (str, bytes)) or not isinstance(allowed_tables, Sequence):
        raise HTTPException(400, "allowed_tables must be a list")
    if any(not isinstance(name, str) for name in allowed_tables):
        raise HTTPException(400, "allowed_tables must contain table names")
    tables = tuple(sorted(set(allowed_tables)))
    registry = _REGISTRY.get(account_id, {})
    if any(name not in registry for name in tables):
        raise HTTPException(400, "one or more allowed tables are unavailable")
    return account_id, tables


def _scope_signatures(key: ScopeKey) -> ScopeSignatures:
    account_id, tables = key
    registry = _REGISTRY.get(account_id, {})
    return tuple((name, _file_sig(registry[name]["location"])) for name in tables)


def _close_scope(key: ScopeKey) -> None:
    with LOCK:
        con = _CONNECTIONS.pop(key, None)
        _SCOPE_SIGNATURES.pop(key, None)
    if con is not None:
        con.close()


def _invalidate_account(account_id: str) -> None:
    with LOCK:
        keys = [scope_key for scope_key in _CONNECTIONS if scope_key[0] == account_id]
        connections = [_CONNECTIONS.pop(key) for key in keys]
        for key in keys:
            _SCOPE_SIGNATURES.pop(key, None)
    for con in connections:
        con.close()


def _evict_account_scopes(account_id: str) -> None:
    with LOCK:
        account_keys = [key for key in _CONNECTIONS if key[0] == account_id]
        evicted: list[duckdb.DuckDBPyConnection] = []
        while len(account_keys) > MAX_SCOPES_PER_ACCOUNT:
            key = account_keys.pop(0)
            con = _CONNECTIONS.pop(key, None)
            _SCOPE_SIGNATURES.pop(key, None)
            if con is not None:
                evicted.append(con)
    for con in evicted:
        con.close()


def _build_scoped_connection(key: ScopeKey, signatures: ScopeSignatures) -> duckdb.DuckDBPyConnection:
    account_id, tables = key
    registry = _REGISTRY.get(account_id, {})
    con = duckdb.connect()
    try:
        _configure_connection(con)
        for name, signature in signatures:
            if signature == "missing":
                raise HTTPException(422, f"dataset {name} could not be reloaded")
            meta = registry[name]
            _load_table(con, name, meta["safe_location"], meta.get("format"))
            meta["rows"] = int(con.execute(f"SELECT count(*) FROM {_qi(name)}").fetchone()[0])
            meta["columns"] = _columns(con, name)
            meta["preview"], meta["preview_truncated"] = _preview(con, name, meta["columns"])
            meta["size_bytes"] = Path(meta["location"]).stat().st_size
            meta["file_signature"] = signature
        if _scope_signatures(key) != signatures:
            raise HTTPException(422, "dataset scope changed while loading")
        con.execute("SET enable_external_access=false")
        return con
    except HTTPException:
        con.close()
        raise
    except Exception as e:  # noqa: BLE001
        con.close()
        LOGGER.warning(
            "dataset_catalog_load_failed account_id=%s error_type=%s",
            account_id,
            type(e).__name__,
        )
        raise HTTPException(422, "dataset scope could not be loaded") from e


def _scoped_connection_for_key(key: ScopeKey) -> duckdb.DuckDBPyConnection:
    signatures = _scope_signatures(key)
    with LOCK:
        con = _CONNECTIONS.get(key)
        if con is not None and _SCOPE_SIGNATURES.get(key) == signatures:
            _CONNECTIONS.move_to_end(key)
            return con
        if con is not None:
            _CONNECTIONS.pop(key, None)
            _SCOPE_SIGNATURES.pop(key, None)
    if con is not None:
        con.close()
    con = _build_scoped_connection(key, signatures)
    with LOCK:
        _CONNECTIONS[key] = con
        _SCOPE_SIGNATURES[key] = signatures
        _CONNECTIONS.move_to_end(key)
    _evict_account_scopes(key[0])
    return con


def _scoped_connection(account_id: str, allowed_tables: Sequence[str]) -> duckdb.DuckDBPyConnection:
    return _scoped_connection_for_key(_scope_key(account_id, allowed_tables))


def inspect_dataset(
    name: str,
    location: str,
    original_name: str,
    expected_format: str | None = None,
) -> dict[str, Any]:
    """Parse a candidate in a one-shot catalog without changing registry state."""
    if name == "schema_version":
        raise HTTPException(400, "reserved table name")
    if not TABLE_RE.fullmatch(name):
        raise HTTPException(400, "invalid table name")
    candidate = Path(location)
    if not candidate.is_file():
        raise HTTPException(404, "dataset file not found")
    safe_location = str(candidate.resolve())
    signature_before = _file_sig(safe_location)
    if signature_before == "missing":
        raise HTTPException(404, "dataset file not found")
    try:
        _validate_expected_file_format(safe_location, expected_format)
    except OSError as exc:
        raise HTTPException(409, "dataset changed while it was being parsed") from exc
    con = duckdb.connect()
    try:
        _configure_connection(con)
        _load_table(con, name, safe_location, expected_format)
        n_rows = con.execute(f"SELECT count(*) FROM {_qi(name)}").fetchone()[0]
        columns = _columns(con, name)
        preview, preview_truncated = _preview(con, name, columns)
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        LOGGER.warning("dataset_parse_failed error_type=%s", type(e).__name__)
        raise HTTPException(422, "dataset could not be parsed") from e
    finally:
        con.close()
    signature_after = _file_sig(safe_location)
    if signature_after == "missing" or signature_after != signature_before:
        raise HTTPException(409, "dataset changed while it was being parsed")
    return {
        "name": name,
        "location": location,
        "safe_location": safe_location,
        "original_name": original_name,
        "rows": int(n_rows),
        "columns": columns,
        "preview": preview,
        "preview_truncated": preview_truncated,
        "size_bytes": int(signature_after.split(":", 1)[0]),
        "file_signature": signature_after,
    }


def register(
    account_id: str,
    name: str,
    location: str,
    kind: str,
    original_name: str,
    url: str | None = None,
    expected_format: str | None = None,
) -> dict[str, Any]:
    if expected_format not in {None, "csv", "json"}:
        raise HTTPException(400, "invalid dataset format")
    safe_location = str(Path(location).resolve())
    signature = _file_sig(safe_location)
    if signature == "missing":
        raise HTTPException(404, "dataset file not found")
    with _account_lock(account_id):
        original = _REGISTRY.get(account_id, {}).get(name)
        if _same_registration(original, safe_location, signature, kind, url, expected_format):
            updated = dict(original)
            updated["original_name"] = original_name
            _REGISTRY[account_id][name] = updated
            return updated

    # Parsing happens before taking the account lock. A slow spreadsheet load
    # therefore blocks neither other accounts nor reads of the previous valid
    # version of this account's dataset.
    prepared = inspect_dataset(name, location, original_name, expected_format)
    safe_location = prepared["safe_location"]
    signature = prepared["file_signature"]
    meta = prepared | {"kind": kind, "url": url, "format": expected_format}
    with _account_lock(account_id):
        _REGISTRY.setdefault(account_id, {})
        current = _REGISTRY[account_id].get(name)
        if current is not original:
            if _same_registration(current, safe_location, signature, kind, url, expected_format):
                return current
            raise HTTPException(409, "dataset changed during registration")
        _REGISTRY[account_id][name] = meta
        _invalidate_account(account_id)
        if original is not None and original["location"] != location:
            return meta | {"previous_location": original["location"]}
        return meta


def current_location(account_id: str, name: str) -> str | None:
    with _account_lock(account_id):
        meta = _REGISTRY.get(account_id, {}).get(name)
        return str(meta["location"]) if meta else None


def begin_preparation(account_id: str, name: str, location: str) -> None:
    safe_location = str(Path(location).resolve())
    with _account_lock(account_id):
        key = (account_id, name, safe_location)
        _PENDING_PREPARATIONS[key] = _PENDING_PREPARATIONS.get(key, 0) + 1


def end_preparation(account_id: str, name: str, location: str) -> None:
    safe_location = str(Path(location).resolve())
    with _account_lock(account_id):
        key = (account_id, name, safe_location)
        remaining = _PENDING_PREPARATIONS.get(key, 1) - 1
        if remaining > 0:
            _PENDING_PREPARATIONS[key] = remaining
        else:
            _PENDING_PREPARATIONS.pop(key, None)


def activate_prepared(
    account_id: str,
    name: str,
    location: str,
    original_name: str,
    url: str,
    expected_format: str,
    expected_previous_location: str | None,
) -> dict[str, Any]:
    """CAS-activate an already prepared immutable connector version."""
    safe_location = str(Path(location).resolve())
    signature = _file_sig(safe_location)
    if signature == "missing":
        raise HTTPException(404, "dataset file not found")
    expected_safe = str(Path(expected_previous_location).resolve()) if expected_previous_location else None
    with _account_lock(account_id):
        original = _REGISTRY.get(account_id, {}).get(name)
        if _same_registration(original, safe_location, signature, "url", url, expected_format):
            result = dict(original)
            result["previous_location"] = expected_previous_location
            return result
        current_safe = str(original["safe_location"]) if original else None
        if current_safe != expected_safe:
            raise HTTPException(409, "dataset changed before connector activation")
        pending_key = (account_id, name, safe_location)
        _PENDING_ACTIVATIONS[pending_key] = _PENDING_ACTIVATIONS.get(pending_key, 0) + 1

    try:
        prepared = inspect_dataset(name, location, original_name, expected_format)
        signature = prepared["file_signature"]
        meta = prepared | {"kind": "url", "url": url, "format": expected_format}
        with _account_lock(account_id):
            current = _REGISTRY.get(account_id, {}).get(name)
            if current is not original:
                if _same_registration(current, safe_location, signature, "url", url, expected_format):
                    result = dict(current)
                    result["previous_location"] = expected_previous_location
                    return result
                raise HTTPException(409, "dataset changed during connector activation")
            _REGISTRY.setdefault(account_id, {})[name] = meta
            _invalidate_account(account_id)
        return meta | {"previous_location": expected_previous_location}
    finally:
        with _account_lock(account_id):
            remaining = _PENDING_ACTIVATIONS.get(pending_key, 1) - 1
            if remaining > 0:
                _PENDING_ACTIVATIONS[pending_key] = remaining
            else:
                _PENDING_ACTIVATIONS.pop(pending_key, None)


def _same_registration(
    meta: dict[str, Any] | None,
    safe_location: str,
    signature: str,
    kind: str,
    url: str | None,
    expected_format: str | None,
) -> bool:
    return bool(
        meta
        and meta.get("safe_location") == safe_location
        and meta.get("file_signature") == signature
        and meta.get("kind") == kind
        and meta.get("url") == url
        and meta.get("format") == expected_format
    )


def _columns(con: duckdb.DuckDBPyConnection, table: str) -> list[dict[str, str]]:
    rows = con.execute(f"DESCRIBE {_qi(table)}").fetchmany(MAX_DATASET_COLUMNS + 1)
    if len(rows) > MAX_DATASET_COLUMNS:
        raise HTTPException(413, "dataset has too many columns")
    columns: list[dict[str, str]] = []
    for row in rows:
        name = str(row[0])
        column_type = str(row[1])
        if len(name) > MAX_DATASET_COLUMN_NAME_CHARS or len(column_type) > MAX_DATASET_TYPE_CHARS:
            raise HTTPException(413, "dataset column metadata exceeds the processing limit")
        columns.append({"name": name, "type": column_type})
    return columns


def _preview_expression(column: dict[str, str]) -> str:
    return _bounded_sql_expression(column["name"], column["type"], MAX_PREVIEW_CELL_CHARS)


def _preview(
    con: duckdb.DuckDBPyConnection,
    table: str,
    columns: Sequence[dict[str, str]],
    n: int = MAX_PREVIEW_ROWS,
) -> tuple[list[list[Any]], bool]:
    if not columns:
        return [], False
    row_limit = min(max(0, n), MAX_PREVIEW_ROWS)
    projection = ", ".join(_preview_expression(column) for column in columns)
    rows = con.execute(f"SELECT {projection} FROM {_qi(table)} LIMIT {row_limit}").fetchall()
    cell_count = max(1, len(rows) * len(columns))
    per_cell_cap = min(MAX_PREVIEW_CELL_CHARS, max(1, MAX_PREVIEW_CHARS // cell_count))
    preview: list[list[Any]] = []
    used_chars = 0
    truncated = False
    for row in rows:
        bounded_row: list[Any] = []
        for raw_value in row:
            value = _json_cell(raw_value)
            value, value_truncated = _bounded_extract_value(value, per_cell_cap)
            rendered_chars = len(str(value)) if value is not None else 0
            remaining = MAX_PREVIEW_CHARS - used_chars
            if rendered_chars > remaining:
                value, _ = _bounded_extract_value(value, max(1, remaining))
                rendered_chars = len(str(value)) if value is not None else 0
                value_truncated = True
            bounded_row.append(value)
            used_chars += rendered_chars
            truncated = truncated or value_truncated
        preview.append(bounded_row)
    return preview, truncated


def list_datasets(account_id: str, *, summary: bool = False) -> list[dict[str, Any]]:
    with _account_lock(account_id):
        meta = _REGISTRY.get(account_id, {})
        if summary:
            return [
                {
                    "table": item["name"],
                    "original_name": item["original_name"],
                    "rows": item["rows"],
                    "exists": Path(item["location"]).exists(),
                }
                for item in meta.values()
            ]
        return [
            {
                "table": item["name"],
                "original_name": item["original_name"],
                "rows": item["rows"],
                "location": item["location"],
                "kind": item["kind"],
                "format": item.get("format"),
                "exists": Path(item["location"]).exists(),
            }
            for item in meta.values()
        ]


def catalog(account_id: str, allowed_tables: Sequence[str]) -> dict[str, Any]:
    """Return only explicitly scoped, prompt-safe schema metadata."""
    with _account_lock(account_id):
        key = _scope_key(account_id, allowed_tables)
        registry = _REGISTRY.get(account_id, {})
        items: list[dict[str, Any]] = []
        used_chars = 0
        omitted = 0
        for name in key[1]:
            meta = registry[name]
            item = {
                "table": name,
                "original_name": meta["original_name"],
                "rows": meta["rows"],
                "columns": meta["columns"],
            }
            item_chars = len(json.dumps(item, ensure_ascii=False, separators=(",", ":"), default=str))
            if used_chars + item_chars > MAX_CATALOG_CHARS:
                omitted += 1
                continue
            items.append(item)
            used_chars += item_chars
    return {
        "datasets": items,
        "total": len(key[1]),
        "returned": len(items),
        "omitted": omitted,
        "truncated": omitted > 0,
    }


def drop(account_id: str, name: str) -> None:
    with _account_lock(account_id):
        if name not in _REGISTRY.get(account_id, {}):
            raise HTTPException(404, f"dataset {name} not found")
        _invalidate_account(account_id)
        del _REGISTRY[account_id][name]
        if not _REGISTRY[account_id]:
            del _REGISTRY[account_id]


def deactivate_if_location(account_id: str, name: str, location: str) -> bool:
    """Drop only the exact active version; stale deferred jobs are harmless."""
    safe_location = str(Path(location).resolve())
    with _account_lock(account_id):
        meta = _REGISTRY.get(account_id, {}).get(name)
        if not meta or meta.get("safe_location") != safe_location:
            return False
        _invalidate_account(account_id)
        del _REGISTRY[account_id][name]
        if not _REGISTRY[account_id]:
            del _REGISTRY[account_id]
        return True


def cleanup_inactive_location(account_id: str, name: str, location: str, cleanup) -> bool:
    with _account_lock(account_id):
        safe_location = str(Path(location).resolve())
        meta = _REGISTRY.get(account_id, {}).get(name)
        if meta and meta.get("safe_location") == safe_location:
            raise HTTPException(409, "active dataset cache versions cannot be deleted")
        if _PENDING_ACTIVATIONS.get((account_id, name, safe_location), 0):
            raise HTTPException(409, "activating dataset cache versions cannot be deleted")
        if _PENDING_PREPARATIONS.get((account_id, name, safe_location), 0):
            raise HTTPException(409, "preparing dataset cache versions cannot be deleted")
        return bool(cleanup(location))


def resync(
    account_id: str,
    name: str,
    fetcher,
    url: str | None = None,
    expected_format: str | None = None,
    cleanup=None,
) -> dict[str, Any]:
    """Re-fetch a URL-based dataset through a fetcher callable -> local path."""
    with _account_lock(account_id):
        meta = _REGISTRY.get(account_id, {}).get(name)
        if not meta:
            raise HTTPException(404, f"dataset {name} not found")
        snapshot = dict(meta)
    location = snapshot["location"]
    remote_url = snapshot.get("url") or url
    kind = snapshot["kind"]
    dataset_format = expected_format or snapshot.get("format")
    if remote_url and dataset_format not in {"csv", "json"}:
        raise HTTPException(400, "format required for URL datasets")
    # Fetchers may perform network I/O. Never hold an account or global lock
    # while calling one.
    if remote_url:
        location = fetcher(remote_url)
        kind = "url"
    try:
        prepared = inspect_dataset(name, location, snapshot["original_name"], dataset_format)
        with _account_lock(account_id):
            current = _REGISTRY.get(account_id, {}).get(name)
            if current is not meta:
                raise HTTPException(409, "dataset changed during refresh")
            updated = prepared | {"kind": kind, "url": remote_url, "format": dataset_format}
            _REGISTRY[account_id][name] = updated
            _invalidate_account(account_id)
    except Exception:
        if cleanup is not None and location != snapshot["location"]:
            _cleanup_quietly(cleanup, location)
        raise
    result = updated | {"table": name}
    if snapshot["location"] != location:
        result["previous_location"] = snapshot["location"]
    return result


def _cleanup_quietly(cleanup, location: str) -> None:
    try:
        cleanup(location)
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("dataset_cache_cleanup_failed error_type=%s", type(exc).__name__)


def _run_with_deadline(con: duckdb.DuckDBPyConnection, operation):
    """Run DuckDB work with an interrupting wall-clock deadline."""
    finished = threading.Event()
    expired = threading.Event()

    def interrupt_after_deadline() -> None:
        if not finished.wait(MAX_DUCKDB_QUERY_SECONDS):
            expired.set()
            try:
                con.interrupt()
            except Exception:  # noqa: BLE001 - the executing thread owns reporting
                pass

    watcher = threading.Thread(target=interrupt_after_deadline, name="duckdb-deadline", daemon=True)
    watcher.start()
    try:
        result = operation()
        if expired.is_set():
            raise HTTPException(504, "query execution timed out")
        return result
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        if expired.is_set():
            raise HTTPException(504, "query execution timed out") from exc
        raise
    finally:
        finished.set()
        watcher.join(timeout=0.1)


def _bounded_sql_expression(name: str, column_type: str, max_chars: int) -> str:
    quoted = _qi(name)
    if _is_bounded_scalar_type(column_type):
        return quoted
    return f"left(CAST({quoted} AS VARCHAR), {max_chars + 1}) AS {quoted}"


def _base_column_type(column_type: str) -> str:
    return column_type.upper().split("(", 1)[0].split("[", 1)[0].strip()


def _is_numeric_column_type(column_type: str) -> bool:
    return _base_column_type(column_type) in {
        "TINYINT",
        "SMALLINT",
        "INTEGER",
        "BIGINT",
        "HUGEINT",
        "UTINYINT",
        "USMALLINT",
        "UINTEGER",
        "UBIGINT",
        "UHUGEINT",
        "DECIMAL",
        "FLOAT",
        "DOUBLE",
        "REAL",
    }


def _is_bounded_scalar_type(column_type: str) -> bool:
    base_type = _base_column_type(column_type)
    return _is_numeric_column_type(column_type) or base_type in {
        "BOOLEAN",
        "DATE",
        "TIME",
        "TIME WITH TIME ZONE",
        "TIMESTAMP",
        "TIMESTAMP WITH TIME ZONE",
        "TIMESTAMP_NS",
        "TIMESTAMP_MS",
        "TIMESTAMP_S",
        "UUID",
    }


def _leading_sql_keyword(sql: str) -> str:
    """Return the first token after whitespace and SQL comments."""
    index = 0
    while index < len(sql):
        if sql[index].isspace():
            index += 1
            continue
        if sql.startswith("--", index):
            newline = sql.find("\n", index + 2)
            index = len(sql) if newline < 0 else newline + 1
            continue
        if sql.startswith("/*", index):
            depth = 1
            index += 2
            while index < len(sql) and depth:
                if sql.startswith("/*", index):
                    depth += 1
                    index += 2
                elif sql.startswith("*/", index):
                    depth -= 1
                    index += 2
                else:
                    index += 1
            continue
        match = re.match(r"[A-Za-z]+", sql[index:])
        return match.group(0).upper() if match else ""
    return ""


def _without_statement_terminator(sql: str) -> str:
    """Remove a parsed statement's optional trailing semicolon, not literal semicolons."""
    normal_semicolons: list[int] = []
    index = 0
    quote: str | None = None
    dollar_quote: str | None = None
    block_depth = 0
    line_comment = False
    while index < len(sql):
        if line_comment:
            if sql[index] == "\n":
                line_comment = False
            index += 1
            continue
        if block_depth:
            if sql.startswith("/*", index):
                block_depth += 1
                index += 2
            elif sql.startswith("*/", index):
                block_depth -= 1
                index += 2
            else:
                index += 1
            continue
        if dollar_quote is not None:
            if sql.startswith(dollar_quote, index):
                index += len(dollar_quote)
                dollar_quote = None
            else:
                index += 1
            continue
        if quote is not None:
            if sql[index] == quote:
                if index + 1 < len(sql) and sql[index + 1] == quote:
                    index += 2
                    continue
                quote = None
            index += 1
            continue
        if sql.startswith("--", index):
            line_comment = True
            index += 2
            continue
        if sql.startswith("/*", index):
            block_depth = 1
            index += 2
            continue
        if sql[index] in {"'", '"'}:
            quote = sql[index]
            index += 1
            continue
        dollar = re.match(r"\$(?:[A-Za-z_][A-Za-z0-9_]*)?\$", sql[index:])
        if dollar:
            dollar_quote = dollar.group(0)
            index += len(dollar_quote)
            continue
        if sql[index] == ";":
            normal_semicolons.append(index)
        index += 1
    if not normal_semicolons:
        return sql
    terminator = normal_semicolons[-1]
    return sql[:terminator] + sql[terminator + 1 :]


def query(account_id: str, sql: str, allowed_tables: Sequence[str]) -> dict[str, Any]:
    sql = sql.strip()
    if not sql:
        raise HTTPException(400, "empty SQL")
    try:
        statements = duckdb.extract_statements(sql)
    except duckdb.Error as exc:
        raise HTTPException(400, "invalid SQL") from exc
    if (
        len(statements) != 1
        or statements[0].type != duckdb.StatementType.SELECT
        or _leading_sql_keyword(sql) not in {"SELECT", "WITH", "VALUES"}
    ):
        raise HTTPException(400, "exactly one read-only query is allowed")
    sql = _without_statement_terminator(sql)
    with _account_lock(account_id):
        key = _scope_key(account_id, allowed_tables)
        try:
            con = _scoped_connection_for_key(key)

            def execute_query():
                schema = con.execute(f"DESCRIBE SELECT * FROM ({sql}\n) AS _q").fetchmany(MAX_QUERY_COLUMNS + 1)
                selected = schema[:MAX_QUERY_COLUMNS]
                if not selected:
                    return schema, [], []
                row_limit = min(MAX_QUERY_ROWS, MAX_QUERY_CELLS // len(selected))
                projection = ", ".join(
                    _bounded_sql_expression(str(column[0]), str(column[1]), MAX_QUERY_CELL_CHARS) for column in selected
                )
                cursor = con.execute(f"SELECT {projection} FROM ({sql}\n) AS _q LIMIT {row_limit + 1}")
                return schema, selected, cursor.fetchmany(row_limit + 1)

            schema, selected, rows = _run_with_deadline(con, execute_query)
        except HTTPException as exc:
            if exc.status_code == 504:
                _close_scope(key)
            raise
        except Exception as e:  # noqa: BLE001
            LOGGER.warning("dataset_query_failed account_id=%s error_type=%s", account_id, type(e).__name__)
            raise HTTPException(422, "query could not be completed") from e

    columns: list[str] = []
    header_chars = 0
    headers_truncated = False
    for column in selected:
        remaining = MAX_QUERY_HEADER_CHARS - header_chars
        if remaining <= 0:
            headers_truncated = True
            break
        name, name_truncated = _bounded_extract_value(str(column[0]), min(MAX_QUERY_COLUMN_NAME_CHARS, remaining))
        assert isinstance(name, str)
        columns.append(name)
        header_chars += len(name)
        headers_truncated = headers_truncated or name_truncated

    column_count = len(columns)
    row_limit = min(MAX_QUERY_ROWS, MAX_QUERY_CELLS // column_count) if column_count else 0
    truncated = headers_truncated or len(schema) > column_count or len(rows) > row_limit
    bounded_rows: list[list[Any]] = []
    used_chars = header_chars
    for row in rows[:row_limit]:
        remaining_chars = MAX_QUERY_CHARS - used_chars
        if remaining_chars <= 0:
            truncated = True
            break
        per_cell_cap = min(MAX_QUERY_CELL_CHARS, max(1, remaining_chars // column_count))
        bounded_row: list[Any] = []
        row_chars = 0
        for raw_value in row[:column_count]:
            value = _json_cell(raw_value)
            value, value_truncated = _bounded_extract_value(value, per_cell_cap)
            row_chars += len(str(value)) if value is not None else 0
            bounded_row.append(value)
            truncated = truncated or value_truncated
        if used_chars + row_chars > MAX_QUERY_CHARS:
            truncated = True
            break
        bounded_rows.append(bounded_row)
        used_chars += row_chars
    return {
        "columns": columns,
        "rows": bounded_rows,
        "row_count": len(bounded_rows),
        "returned_row_count": len(bounded_rows),
        "columns_truncated": len(schema) > len(columns) or headers_truncated,
        "truncated": truncated,
    }


def describe(account_id: str, table: str, allowed_tables: Sequence[str]) -> dict[str, Any]:
    with _account_lock(account_id):
        key = _scope_key(account_id, allowed_tables)
        if table not in key[1]:
            raise HTTPException(400, f"dataset {table} is not allowed in this scope")
        meta = _REGISTRY[account_id][table]
        con = _scoped_connection_for_key(key)

        def execute_describe() -> dict[str, Any]:
            selected_columns = meta["columns"][:MAX_DESCRIBE_COLUMNS]
            projection = ", ".join(
                _bounded_sql_expression(column["name"], column["type"], MAX_DESCRIBE_VALUE_CHARS)
                for column in selected_columns
            )
            sampled_sql = (
                f"SELECT {projection} FROM {_qi(table)} LIMIT {MAX_DESCRIBE_ROWS}"
                if projection
                else f"SELECT * FROM {_qi(table)} LIMIT {MAX_DESCRIBE_ROWS}"
            )
            cursor = con.execute(f"SUMMARIZE SELECT * FROM ({sampled_sql}) AS _sample")
            summary_columns = [str(column[0]) for column in cursor.description]
            summaries = [dict(zip(summary_columns, row, strict=True)) for row in cursor.fetchall()]
            stats: dict[str, Any] = {
                "table": table,
                "rows": meta["rows"],
                "profiled_rows": min(meta["rows"], MAX_DESCRIBE_ROWS),
                "columns_truncated": len(meta["columns"]) > MAX_DESCRIBE_COLUMNS,
                "columns": [],
            }
            remaining_chars = MAX_DESCRIBE_CHARS - len(table)
            response_truncated = bool(stats["columns_truncated"])

            def bounded_text(value: Any, max_chars: int = MAX_DESCRIBE_VALUE_CHARS) -> str | None:
                nonlocal remaining_chars, response_truncated
                if value is None:
                    return None
                if remaining_chars <= 0:
                    response_truncated = True
                    return None
                bounded, was_truncated = _bounded_extract_value(str(value), min(max_chars, remaining_chars))
                assert isinstance(bounded, str)
                remaining_chars -= len(bounded)
                response_truncated = response_truncated or was_truncated
                return bounded

            for column_index, (column, summary) in enumerate(zip(selected_columns, summaries, strict=False)):
                cname = column["name"]
                ctype = column["type"]
                # Names/types were strictly bounded at registration, and the
                # selected-column count ensures these required fields fit.
                remaining_chars -= len(cname) + len(ctype)
                if remaining_chars < 0:
                    response_truncated = True
                entry: dict[str, Any] = {"name": cname, "type": ctype}
                try:
                    if _is_numeric_column_type(ctype):
                        entry.update(
                            {
                                "min": _summary_numeric_cell(summary["min"], ctype),
                                "max": _summary_numeric_cell(summary["max"], ctype),
                                "avg": _summary_numeric_cell(summary["avg"], ctype),
                                "distinct": int(summary["approx_unique"] or 0),
                            }
                        )
                    elif _base_column_type(ctype) in {
                        "DATE",
                        "TIME",
                        "TIME WITH TIME ZONE",
                        "TIMESTAMP",
                        "TIMESTAMP WITH TIME ZONE",
                        "TIMESTAMP_NS",
                        "TIMESTAMP_MS",
                        "TIMESTAMP_S",
                    }:
                        entry.update(
                            {
                                "min": bounded_text(summary["min"]),
                                "max": bounded_text(summary["max"]),
                                "distinct": int(summary["approx_unique"] or 0),
                            }
                        )
                    else:
                        entry["distinct"] = int(summary["approx_unique"] or 0)
                        if column_index < MAX_TOP_VALUE_COLUMNS and remaining_chars > 0:
                            top = con.execute(
                                f"SELECT {_qi(cname)}, count(*) AS n FROM ({sampled_sql}) AS _sample "
                                "GROUP BY 1 ORDER BY n DESC LIMIT 6"
                            ).fetchall()
                            top_values: list[dict[str, Any]] = []
                            for colval, count in top:
                                value = bounded_text(colval)
                                if value is None and colval is not None:
                                    break
                                top_values.append({"value": value, "count": int(count)})
                            entry["top_values"] = top_values
                        elif column_index < MAX_TOP_VALUE_COLUMNS:
                            response_truncated = True
                except (KeyError, TypeError, ValueError):
                    pass
                stats["columns"].append(entry)
            stats["truncated"] = response_truncated
            return stats

        try:
            return _run_with_deadline(con, execute_describe)
        except HTTPException as exc:
            if exc.status_code == 504:
                _close_scope(key)
            raise
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("dataset_describe_failed account_id=%s error_type=%s", account_id, type(exc).__name__)
            raise HTTPException(422, "dataset could not be described") from exc


def _extract_loaded(
    con: duckdb.DuckDBPyConnection,
    table: str,
    meta: dict[str, Any],
    max_rows: int,
) -> dict[str, Any]:
    if not isinstance(max_rows, int) or isinstance(max_rows, bool) or not 1 <= max_rows <= MAX_EXTRACT_ROWS:
        raise HTTPException(400, f"max_rows must be between 1 and {MAX_EXTRACT_ROWS}")
    all_columns = [str(column["name"]) for column in meta["columns"]]
    columns: list[str] = []
    header_chars = 0
    headers_truncated = False
    for raw_name in all_columns[:MAX_EXTRACT_COLUMNS]:
        remaining_header_chars = MAX_EXTRACT_HEADER_CHARS - header_chars
        if remaining_header_chars <= 0:
            headers_truncated = True
            break
        name, name_truncated = _bounded_extract_value(
            raw_name,
            min(MAX_EXTRACT_COLUMN_NAME_CHARS, remaining_header_chars),
        )
        assert isinstance(name, str)
        columns.append(name)
        header_chars += len(name)
        headers_truncated = headers_truncated or name_truncated
    columns_truncated = headers_truncated or len(columns) < len(all_columns)
    if not columns:
        raise HTTPException(422, "dataset has no extractable columns")
    row_limit = min(max_rows, MAX_EXTRACT_CELLS // len(columns))
    selected_meta = meta["columns"][: len(columns)]
    selected_sql = ", ".join(
        _bounded_sql_expression(column["name"], column["type"], MAX_EXTRACT_CELL_CHARS) for column in selected_meta
    )
    rows = _run_with_deadline(
        con,
        lambda: con.execute(f"SELECT {selected_sql} FROM {_qi(table)} LIMIT {row_limit + 1}").fetchall(),
    )
    total_row_count = int(meta["rows"])
    truncated = columns_truncated or len(rows) > row_limit or row_limit < min(max_rows, total_row_count)
    bounded: list[list[Any]] = []
    used_chars = header_chars
    column_count = len(columns)
    for row in rows[:row_limit]:
        normalized = [_json_cell(value) for value in row]
        per_cell_cap = min(MAX_EXTRACT_CELL_CHARS, max(32, (MAX_EXTRACT_CHARS - used_chars) // column_count))
        bounded_row: list[Any] = []
        row_chars = 0
        for value in normalized:
            value, value_truncated = _bounded_extract_value(value, per_cell_cap)
            truncated = truncated or value_truncated
            row_chars += len(str(value)) if value is not None else 0
            bounded_row.append(value)
        if used_chars + row_chars > MAX_EXTRACT_CHARS:
            truncated = True
            break
        bounded.append(bounded_row)
        used_chars += row_chars
        if used_chars >= MAX_EXTRACT_CHARS:
            truncated = True
            break
    return {
        "columns": columns,
        "rows": bounded,
        "row_count": total_row_count,
        "total_row_count": total_row_count,
        "returned_row_count": len(bounded),
        "columns_truncated": columns_truncated,
        "truncated": truncated,
    }


def extract(
    account_id: str,
    table: str,
    allowed_tables: Sequence[str],
    max_rows: int = 500,
) -> dict[str, Any]:
    """Return a bounded snapshot for ingestion without loading a file in Node."""
    with _account_lock(account_id):
        key = _scope_key(account_id, allowed_tables)
        if table not in key[1]:
            raise HTTPException(400, "dataset is not allowed in this scope")
        con = _scoped_connection_for_key(key)
        meta = _REGISTRY[account_id][table]
        try:
            return _extract_loaded(con, table, meta, max_rows)
        except HTTPException as exc:
            if exc.status_code == 504:
                _close_scope(key)
            raise


def extract_candidate(
    name: str,
    location: str,
    expected_format: str,
    max_rows: int = 500,
) -> dict[str, Any]:
    """Extract a proven prepared cache file without exposing it in the registry."""
    meta = inspect_dataset(name, location, Path(location).name, expected_format)
    con = duckdb.connect()
    try:
        _configure_connection(con)
        _load_table(con, name, meta["safe_location"], expected_format)
        return _extract_loaded(con, name, meta, max_rows)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("dataset_candidate_extract_failed error_type=%s", type(exc).__name__)
        raise HTTPException(422, "dataset candidate could not be extracted") from exc
    finally:
        con.close()


def _decimal_string(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return "0" if text in {"-0", ""} else text


def _bounded_extract_value(value: Any, max_chars: int) -> tuple[Any, bool]:
    if isinstance(value, (list, dict)):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)
    if isinstance(value, str) and len(value) > max_chars:
        return value[: max(0, max_chars - 1)] + "…", True
    return value, False


def _json_cell(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, Decimal):
        return _decimal_string(value) if value.is_finite() else None
    if isinstance(value, dict):
        return {str(key): _json_cell(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_cell(item) for item in value]
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, int) and not isinstance(value, bool) and abs(value) > JS_MAX_SAFE_INTEGER:
        return str(value)
    return value


def _summary_numeric_cell(value: Any, column_type: str) -> float | str | None:
    if value is None:
        return None
    if "DECIMAL" in column_type:
        try:
            decimal = value if isinstance(value, Decimal) else Decimal(str(value))
            return _decimal_string(decimal) if decimal.is_finite() else None
        except Exception:  # noqa: BLE001
            return None
    if "INT" in column_type:
        try:
            decimal = value if isinstance(value, Decimal) else Decimal(str(value))
            if not decimal.is_finite():
                return None
            if decimal == decimal.to_integral_value():
                integer = int(decimal)
                normalized = _json_cell(integer)
                return float(normalized) if isinstance(normalized, int) else normalized
            return float(decimal)
        except Exception:  # noqa: BLE001
            return None
    try:
        number = float(value) if value is not None else None
        return number if number is None or math.isfinite(number) else None
    except (TypeError, ValueError):
        return None

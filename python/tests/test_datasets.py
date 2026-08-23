"""Characterization tests for the DuckDB dataset registry and SQL validator.

These pin the current behaviour of the read-only SQL guard and the
result-normalization in ``datasets.query``. Use a distinct account_id per test —
the module-level ``_REGISTRY`` persists across tests in-process.
"""

from __future__ import annotations

import time
import uuid
from concurrent.futures import ThreadPoolExecutor
from decimal import Decimal
from pathlib import Path
from threading import Barrier, Event

import duckdb
import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from app import datasets


def _acct() -> str:
    return f"acct-{uuid.uuid4().hex[:12]}"


def _register_csv(tmp_path, account: str, name: str, csv_text: str):
    p = tmp_path / f"{account}_{name}.csv"
    p.write_text(csv_text)
    return datasets.register(account, name, str(p), "path", f"{name}.csv", None)


def _cached_connection(account: str, tables: list[str]) -> duckdb.DuckDBPyConnection:
    key = (account, tuple(sorted(set(tables))))
    with datasets.LOCK:
        return datasets._CONNECTIONS[key]


def _assert_closed(con: duckdb.DuckDBPyConnection) -> None:
    with datasets.LOCK:
        with pytest.raises(duckdb.ConnectionException, match="already closed"):
            con.execute("SELECT 1")


def test_european_csv_detector_is_conservative():
    assert datasets._is_european_semicolon_csv("Date;Amount;Text\n18.08.26;-1.234,56;Rent\n")
    assert not datasets._is_european_semicolon_csv("date,amount\n2026-08-18,12.50\n")
    assert not datasets._is_european_semicolon_csv("note;18.08.26\n")


def test_existing_sample_csvs_keep_default_reader():
    for sample in sorted((datasets.REPO_ROOT / "data" / "sample").glob("*.csv")):
        assert datasets._read_sql(str(sample)) == "read_csv_auto(?)"


def test_registers_sparkasse_csv_with_dates_and_decimal_amounts(tmp_path):
    account = _acct()
    csv_text = (
        "\ufeffKontoinhaber:;Example User\n"
        "Erstellt am:;22.08.2026\n"
        "\n"
        "Buchungsdatum;Wertstellung;Status;Betrag (€)\n"
        '18.08.26;18.08.26;Gebucht;"-1.234,56"\n'
        '19.08.26;19.08.26;Gebucht;"42,10"\n'
    )

    meta = _register_csv(tmp_path, account, "sparkasse", csv_text)
    columns = {column["name"]: column["type"] for column in meta["columns"]}
    assert columns["Buchungsdatum"] == "DATE"
    assert columns["Betrag (€)"].startswith("DECIMAL")

    result = datasets.query(
        account,
        'SELECT min("Buchungsdatum")::VARCHAR AS first_date, sum("Betrag (€)") AS amount FROM sparkasse',
        ["sparkasse"],
    )
    assert result["rows"][0][0] == "2026-08-18"
    assert result["rows"][0][1] == "-1192.46"


# ---------------------------------------------------------------------------
# SQL validator contract
# ---------------------------------------------------------------------------
@pytest.mark.parametrize(
    "sql",
    [
        "INSERT INTO x VALUES (1)",
        "UPDATE t SET a=1",
        "DELETE FROM t",
        "DROP TABLE t",
        "ALTER TABLE t ADD COLUMN a int",
        "CREATE TABLE t (a int)",
        "ATTACH 'x'",
        "COPY t TO 'a'",
        "CALL some_fn()",
        "INSTALL foo",
        "LOAD foo",
        "SELECT 1; DROP TABLE t",
        "",
    ],
)
def test_query_rejects_mutations_and_ddl(sql: str):
    with pytest.raises(HTTPException) as exc:
        datasets.query(_acct(), sql, [])
    assert exc.value.status_code == 400


def test_query_accepts_read_statements():
    account = _acct()
    r = datasets.query(account, "SELECT 1 AS n", [])
    assert r["columns"] == ["n"]
    assert r["rows"] == [[1]]
    assert r["row_count"] == 1
    with datasets.LOCK:
        assert (account, ()) in datasets._CONNECTIONS


def test_query_accepts_leading_whitespace():
    r = datasets.query(_acct(), "  SELECT 2 AS n", [])
    assert r["rows"] == [[2]]


@pytest.mark.parametrize(
    "sql",
    [
        "PRAGMA functions",
        "/* leading comment */ PRAGMA threads=100000",
        "PRAGMA memory_limit='1TB'",
        "SELECT 1; SELECT 2",
        "SELECT 1; PRAGMA threads=1",
    ],
)
def test_query_rejects_pragma_and_multiple_statements(sql: str):
    with pytest.raises(HTTPException) as error:
        datasets.query(_acct(), sql, [])

    assert error.value.status_code == 400


@pytest.mark.parametrize("sql", ["SELECT 1 AS n;", "SELECT ';' AS n; -- trailing comment"])
def test_query_accepts_one_select_with_a_trailing_terminator(sql: str):
    assert datasets.query(_acct(), sql, [])["row_count"] == 1


# ---------------------------------------------------------------------------
# Result normalization
# ---------------------------------------------------------------------------
def test_query_preserves_large_int_cells_as_exact_strings(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "big", "id\n1234567890123456789\n")
    r = datasets.query(account, "SELECT id FROM big", ["big"])
    cell = r["rows"][0][0]
    assert cell == "1234567890123456789"


def test_query_preserves_decimal_cells_as_exact_strings():
    result = datasets.query(_acct(), "SELECT 123456789012345678.123456::DECIMAL(36, 6) AS amount", [])

    assert result["rows"] == [["123456789012345678.123456"]]


def test_query_caps_results_at_500_rows(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "many", "x\n" + "".join(f"{i}\n" for i in range(1000)))
    r = datasets.query(account, "SELECT * FROM many", ["many"])
    assert r["row_count"] == 500
    assert len(r["rows"]) == 500


def test_query_bounds_columns_cells_and_strings(monkeypatch):
    monkeypatch.setattr(datasets, "MAX_QUERY_COLUMNS", 2)
    monkeypatch.setattr(datasets, "MAX_QUERY_CELLS", 4)
    monkeypatch.setattr(datasets, "MAX_QUERY_CELL_CHARS", 5)

    result = datasets.query(
        _acct(),
        "SELECT repeat('x', 100) AS text, i, i + 1 AS omitted FROM range(5) AS rows(i)",
        [],
    )

    assert result["columns"] == ["text", "i"]
    assert result["rows"] == [["xxxx…", 0], ["xxxx…", 1]]
    assert result["returned_row_count"] == 2
    assert result["columns_truncated"] is True
    assert result["truncated"] is True


def test_query_bounds_nested_values_inside_duckdb(monkeypatch):
    monkeypatch.setattr(datasets, "MAX_QUERY_CELL_CHARS", 20)

    result = datasets.query(_acct(), "SELECT {'items': [repeat('x', 10000)]} AS nested", [])

    assert isinstance(result["rows"][0][0], str)
    assert len(result["rows"][0][0]) == 20
    assert result["rows"][0][0].endswith("…")
    assert result["truncated"] is True


def test_query_deadline_interrupts_duckdb_at_the_lowest_boundary(monkeypatch):
    interrupted = Event()

    class BlockingConnection:
        def execute(self, _sql):
            assert interrupted.wait(timeout=1)
            raise RuntimeError("interrupted")

        def interrupt(self):
            interrupted.set()

    monkeypatch.setattr(datasets, "MAX_DUCKDB_QUERY_SECONDS", 0.01)
    monkeypatch.setattr(datasets, "_scoped_connection_for_key", lambda _key: BlockingConnection())
    started = time.monotonic()

    with pytest.raises(HTTPException) as error:
        datasets.query(_acct(), "SELECT 1", [])

    assert error.value.status_code == 504
    assert interrupted.is_set()
    assert time.monotonic() - started < 0.5


def test_describe_interrupt_is_not_swallowed_by_top_value_profiling(monkeypatch):
    account = _acct()
    interrupted = Event()
    executed: list[str] = []

    class SummaryCursor:
        description = [
            ("column_name",),
            ("column_type",),
            ("min",),
            ("max",),
            ("avg",),
            ("approx_unique",),
        ]

        @staticmethod
        def fetchall():
            return [("text", "VARCHAR", "a", "z", None, 2)]

    class BlockingDescribeConnection:
        def execute(self, sql):
            executed.append(sql)
            if sql.startswith("SUMMARIZE"):
                return SummaryCursor()
            assert interrupted.wait(timeout=1)
            raise RuntimeError("interrupted")

        def interrupt(self):
            interrupted.set()

    with datasets._account_lock(account):
        datasets._REGISTRY[account] = {
            "text_data": {
                "name": "text_data",
                "rows": 2,
                "columns": [{"name": "text", "type": "VARCHAR"}],
                "location": "/nonexistent",
            }
        }
    monkeypatch.setattr(datasets, "MAX_DUCKDB_QUERY_SECONDS", 0.01)
    monkeypatch.setattr(datasets, "_scoped_connection_for_key", lambda _key: BlockingDescribeConnection())
    try:
        with pytest.raises(HTTPException) as error:
            datasets.describe(account, "text_data", ["text_data"])
    finally:
        with datasets._account_lock(account):
            datasets._REGISTRY.pop(account, None)

    assert error.value.status_code == 504
    assert interrupted.is_set()
    assert len(executed) == 2


def test_query_normalizes_infinity_to_none():
    result = datasets.query(_acct(), "SELECT 1.0 / 0.0 AS x", [])

    assert result["rows"] == [[None]]


def test_json_cell_normalizes_nonfinite_decimals_to_none():
    assert datasets._json_cell(Decimal("NaN")) is None
    assert datasets._json_cell(Decimal("Infinity")) is None


def test_json_cell_preserves_nested_decimals_and_binary_values():
    assert datasets._json_cell({"amounts": [Decimal("1.2300")], "raw": b"\x00\xff"}) == {
        "amounts": ["1.23"],
        "raw": "00ff",
    }


# ---------------------------------------------------------------------------
# Persistent connection lifecycle and account isolation
# ---------------------------------------------------------------------------
def test_repeated_queries_reload_only_when_file_changes(tmp_path, monkeypatch):
    account = _acct()
    original_read_sql = datasets._read_sql
    calls: list[str] = []

    def _tracked_read_sql(path: str, expected_format: str | None = None) -> str:
        calls.append(path)
        return original_read_sql(path, expected_format)

    monkeypatch.setattr(datasets, "_read_sql", _tracked_read_sql)
    _register_csv(tmp_path, account, "memo", "value\n1\n")

    assert datasets.query(account, "SELECT value FROM memo", ["memo"])["rows"] == [[1]]
    calls_after_first_query = len(calls)
    assert calls_after_first_query == 2
    first_connection = _cached_connection(account, ["memo"])
    assert datasets.query(account, "SELECT value FROM memo", ["memo"])["rows"] == [[1]]
    assert len(calls) == calls_after_first_query

    path = tmp_path / f"{account}_memo.csv"
    path.write_text("value\n200\n300\n")
    assert datasets.query(account, "SELECT value FROM memo ORDER BY value", ["memo"])["rows"] == [[200], [300]]
    assert len(calls) == calls_after_first_query + 1
    assert _cached_connection(account, ["memo"]) is not first_connection
    _assert_closed(first_connection)


def test_missing_file_fails_closed_for_query_and_describe(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "missing", "value\n1\n")
    assert datasets.query(account, "SELECT * FROM missing", ["missing"])["rows"] == [[1]]

    (tmp_path / f"{account}_missing.csv").unlink()

    with pytest.raises(HTTPException) as query_error:
        datasets.query(account, "SELECT * FROM missing", ["missing"])
    assert query_error.value.status_code == 422

    with pytest.raises(HTTPException) as describe_error:
        datasets.describe(account, "missing", ["missing"])
    assert describe_error.value.status_code == 422

    path = tmp_path / f"{account}_missing.csv"
    path.write_text("value\n9\n10\n")
    assert datasets.query(account, "SELECT * FROM missing ORDER BY value", ["missing"])["rows"] == [[9], [10]]
    assert datasets.describe(account, "missing", ["missing"])["rows"] == 2


def test_account_catalogs_are_isolated(tmp_path):
    account_a = _acct()
    account_b = _acct()
    _register_csv(tmp_path, account_a, "shared", "canary\nalpha\n")
    _register_csv(tmp_path, account_b, "shared", "canary\nbeta\n")
    _register_csv(tmp_path, account_a, "only_a", "secret\naccount-a-only\n")

    for _ in range(2):
        assert datasets.query(account_a, "SELECT canary FROM shared", ["shared"])["rows"] == [["alpha"]]
        assert datasets.query(account_b, "SELECT canary FROM shared", ["shared"])["rows"] == [["beta"]]

    with pytest.raises(HTTPException) as error:
        datasets.query(account_b, "SELECT * FROM only_a", ["shared"])
    assert error.value.status_code == 422


def test_concurrent_queries_keep_account_catalogs_isolated(tmp_path):
    account_a = _acct()
    account_b = _acct()
    _register_csv(tmp_path, account_a, "shared", "canary\nalpha\n")
    _register_csv(tmp_path, account_b, "shared", "canary\nbeta\n")

    def _canary(account: str) -> str:
        return datasets.query(account, "SELECT canary FROM shared", ["shared"])["rows"][0][0]

    accounts = [account_a, account_b] * 10
    with ThreadPoolExecutor(max_workers=4) as executor:
        results = list(executor.map(_canary, accounts))

    assert results == ["alpha", "beta"] * 10


def test_describe_quotes_column_identifiers(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "quoted", '"we""ird"\n1\n2\n')

    result = datasets.describe(account, "quoted", ["quoted"])

    assert result["columns"][0]["name"] == 'we"ird'
    assert result["columns"][0]["min"] == 1.0
    assert result["columns"][0]["max"] == 2.0


# ---------------------------------------------------------------------------
# Scoped catalog enforcement and bounded cache
# ---------------------------------------------------------------------------
def test_scoped_catalog_exposes_only_allowed_tables(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "alpha", "canary\na\n")
    _register_csv(tmp_path, account, "beta", "canary\nb\n")

    assert datasets.query(account, "SELECT canary FROM alpha", ["alpha", "alpha"])["rows"] == [["a"]]
    with pytest.raises(HTTPException) as unselected_error:
        datasets.query(account, "SELECT canary FROM beta", ["alpha"])
    assert unselected_error.value.status_code == 422

    with pytest.raises(HTTPException) as empty_error:
        datasets.query(account, "SELECT canary FROM alpha", [])
    assert empty_error.value.status_code == 422

    visible = datasets.query(
        account,
        "SELECT table_name FROM duckdb_tables() WHERE NOT internal ORDER BY table_name",
        ["beta", "alpha", "alpha"],
    )
    assert visible["rows"] == [["alpha"], ["beta"]]


def test_scope_rejects_unregistered_allowed_table(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "alpha", "x\n1\n")

    with pytest.raises(HTTPException) as error:
        datasets.query(account, "SELECT 1", ["missing"])

    assert error.value.status_code == 400
    assert error.value.detail == "one or more allowed tables are unavailable"


def test_describe_rejects_table_outside_allowlist(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "alpha", "x\n1\n")
    _register_csv(tmp_path, account, "beta", "x\n2\n")

    with pytest.raises(HTTPException) as error:
        datasets.describe(account, "beta", ["alpha"])

    assert error.value.status_code == 400


@pytest.mark.parametrize(
    ("reader", "filename", "content"),
    [
        ("read_csv_auto", "outside.csv", "x\n1\n"),
        ("read_json_auto", "outside.json", '[{"x":1}]'),
        ("read_parquet", "outside.parquet", "not-a-parquet-file"),
    ],
)
def test_scoped_catalog_disables_external_table_functions(tmp_path, reader: str, filename: str, content: str):
    account = _acct()
    outside = tmp_path / filename
    outside.write_text(content)
    escaped_path = str(outside).replace("'", "''")

    setting = datasets.query(account, "SELECT current_setting('enable_external_access') AS enabled", [])
    assert setting["rows"] == [[False]]

    with pytest.raises(HTTPException) as error:
        datasets.query(account, f"SELECT * FROM {reader}('{escaped_path}')", [])
    assert error.value.status_code == 422
    assert error.value.detail == "query could not be completed"


def test_scoped_catalog_cannot_reenable_external_access():
    account = _acct()
    assert datasets.query(account, "SELECT 1", [])["rows"] == [[1]]

    with pytest.raises(HTTPException) as error:
        datasets.query(account, "PRAGMA enable_external_access=true", [])
    assert error.value.status_code == 400
    setting = datasets.query(account, "SELECT current_setting('enable_external_access') AS enabled", [])
    assert setting["rows"] == [[False]]


def test_register_invalidates_every_cached_account_scope(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "alpha", "x\n1\n")
    _register_csv(tmp_path, account, "beta", "x\n2\n")
    datasets.query(account, "SELECT * FROM alpha", ["alpha"])
    datasets.query(account, "SELECT * FROM beta", ["beta"])
    alpha_connection = _cached_connection(account, ["alpha"])
    beta_connection = _cached_connection(account, ["beta"])

    _register_csv(tmp_path, account, "gamma", "x\n3\n")

    with datasets.LOCK:
        assert not any(key[0] == account for key in datasets._CONNECTIONS)
    _assert_closed(alpha_connection)
    _assert_closed(beta_connection)


def test_resync_invalidates_cached_account_scopes(tmp_path):
    account = _acct()
    initial = tmp_path / f"{account}_remote_initial.csv"
    initial.write_text("x\n1\n")
    datasets.register(account, "remote", str(initial), "url", "remote.csv", "https://example.test/data", "csv")
    assert datasets.query(account, "SELECT x FROM remote", ["remote"])["rows"] == [[1]]
    old_connection = _cached_connection(account, ["remote"])

    refreshed = tmp_path / f"{account}_remote_refreshed.csv"
    refreshed.write_text("x\n2\n3\n")
    datasets.resync(account, "remote", fetcher=lambda _url: str(refreshed), expected_format="csv")

    _assert_closed(old_connection)
    assert datasets.query(account, "SELECT x FROM remote ORDER BY x", ["remote"])["rows"] == [[2], [3]]


def test_resync_upgrades_legacy_path_metadata_when_connector_url_is_supplied(tmp_path):
    account = _acct()
    initial = tmp_path / f"{account}_legacy.csv"
    initial.write_text("x\n1\n")
    datasets.register(account, "remote", str(initial), "path", "Remote feed", None)

    refreshed = tmp_path / f"{account}_refreshed.json"
    refreshed.write_text('[{"x": 2}]')
    remote_url = "https://example.test/data.json"
    result = datasets.resync(
        account,
        "remote",
        fetcher=lambda url: str(refreshed) if url == remote_url else "",
        url=remote_url,
        expected_format="json",
    )

    assert result["kind"] == "url"
    assert result["url"] == remote_url
    assert result["location"] == str(refreshed)
    assert datasets.query(account, "SELECT x FROM remote", ["remote"])["rows"] == [[2]]


def test_drop_invalidates_every_cached_account_scope(tmp_path):
    account = _acct()
    _register_csv(tmp_path, account, "alpha", "x\n1\n")
    _register_csv(tmp_path, account, "beta", "x\n2\n")
    datasets.query(account, "SELECT * FROM alpha", ["alpha"])
    datasets.query(account, "SELECT * FROM beta", ["beta"])
    alpha_connection = _cached_connection(account, ["alpha"])
    beta_connection = _cached_connection(account, ["beta"])

    datasets.drop(account, "alpha")

    _assert_closed(alpha_connection)
    _assert_closed(beta_connection)
    assert datasets.query(account, "SELECT * FROM beta", ["beta"])["rows"] == [[2]]


def test_scope_lru_evicts_only_within_account(tmp_path):
    account = _acct()
    other_account = _acct()
    for index in range(4):
        _register_csv(tmp_path, account, f"table_{index}", f"x\n{index}\n")
    _register_csv(tmp_path, other_account, "table_0", "x\n99\n")
    datasets.query(other_account, "SELECT * FROM table_0", ["table_0"])
    other_connection = _cached_connection(other_account, ["table_0"])

    scopes = [
        ["table_0"],
        ["table_1"],
        ["table_2"],
        ["table_3"],
        ["table_0", "table_1"],
        ["table_0", "table_2"],
        ["table_0", "table_3"],
        ["table_1", "table_2"],
        ["table_1", "table_3"],
    ]
    first_connection = None
    second_connection = None
    for index, scope in enumerate(scopes[: datasets.MAX_SCOPES_PER_ACCOUNT]):
        datasets.query(account, "SELECT 1", scope)
        if index == 0:
            first_connection = _cached_connection(account, scope)
        elif index == 1:
            second_connection = _cached_connection(account, scope)

    datasets.query(account, "SELECT 1", scopes[0])
    datasets.query(account, "SELECT 1", scopes[-1])

    with datasets.LOCK:
        account_keys = [key for key in datasets._CONNECTIONS if key[0] == account]
        assert len(account_keys) == datasets.MAX_SCOPES_PER_ACCOUNT
        assert (account, ("table_0",)) in datasets._CONNECTIONS
        assert (account, ("table_1",)) not in datasets._CONNECTIONS
        assert (other_account, ("table_0",)) in datasets._CONNECTIONS
    assert first_connection is not None
    assert second_connection is not None
    with datasets.LOCK:
        assert first_connection.execute("SELECT x FROM table_0").fetchall() == [(0,)]
    _assert_closed(second_connection)
    with datasets.LOCK:
        assert other_connection.execute("SELECT x FROM table_0").fetchall() == [(99,)]


# ---------------------------------------------------------------------------
# Register / drop lifecycle (write paths use one-shot validation catalogs)
# ---------------------------------------------------------------------------
def test_register_and_drop_lifecycle(tmp_path):
    account = _acct()
    a = tmp_path / "a.csv"
    a.write_text("x\n1\n2\n")
    b = tmp_path / "b.csv"
    b.write_text("y\n3\n")
    datasets.register(account, "a", str(a), "path", "a.csv", None)
    datasets.register(account, "b", str(b), "path", "b.csv", None)

    assert datasets.query(account, "SELECT count(*) AS n FROM a", ["a"])["rows"][0][0] == 2

    datasets.drop(account, "a")
    with pytest.raises(HTTPException):
        datasets.query(account, "SELECT * FROM a", [])

    assert datasets.query(account, "SELECT count(*) AS n FROM b", ["b"])["rows"][0][0] == 1
    datasets.drop(account, "b")
    with datasets.LOCK:
        assert not any(key[0] == account for key in datasets._CONNECTIONS)


def test_register_drop_do_not_use_scoped_query_catalog(tmp_path, monkeypatch):
    def _boom(account_id, allowed_tables):
        raise AssertionError("_scoped_connection must not be called on write paths")

    monkeypatch.setattr(datasets, "_scoped_connection", _boom)
    account = _acct()
    a = tmp_path / "a.csv"
    a.write_text("x\n1\n")
    datasets.register(account, "a", str(a), "path", "a.csv", None)
    datasets.drop(account, "a")


def test_register_is_signature_aware_and_idempotent(tmp_path, monkeypatch):
    account = _acct()
    source = tmp_path / "stable.csv"
    source.write_text("value\n1\n")
    original_inspect = datasets.inspect_dataset
    inspected: list[str] = []

    def tracked_inspect(*args, **kwargs):
        inspected.append(str(args[1]))
        return original_inspect(*args, **kwargs)

    monkeypatch.setattr(datasets, "inspect_dataset", tracked_inspect)
    first = datasets.register(account, "stable", str(source), "path", "first.csv")
    assert datasets.query(account, "SELECT value FROM stable", ["stable"])["rows"] == [[1]]
    connection = _cached_connection(account, ["stable"])

    second = datasets.register(account, "stable", str(source), "path", "renamed.csv")

    assert inspected == [str(source)]
    assert second["original_name"] == "renamed.csv"
    assert second["file_signature"] == first["file_signature"]
    assert _cached_connection(account, ["stable"]) is connection

    source.write_text("value\n2\n3\n")
    third = datasets.register(account, "stable", str(source), "path", "renamed.csv")

    assert inspected == [str(source), str(source)]
    assert third["file_signature"] != first["file_signature"]
    _assert_closed(connection)
    assert datasets.query(account, "SELECT value FROM stable ORDER BY value", ["stable"])["rows"] == [[2], [3]]


def test_register_rejects_a_file_that_changes_during_parse(tmp_path, monkeypatch):
    account = _acct()
    source = tmp_path / "moving.csv"
    source.write_text("value\n1\n")
    original_load = datasets._load_table

    def moving_load(*args, **kwargs):
        original_load(*args, **kwargs)
        source.write_text("value\n2\n3\n")

    monkeypatch.setattr(datasets, "_load_table", moving_load)

    with pytest.raises(HTTPException) as error:
        datasets.register(account, "moving", str(source), "path", "moving.csv")

    assert error.value.status_code == 409
    assert "moving" not in datasets._REGISTRY.get(account, {})


def test_concurrent_registration_uses_compare_and_swap(tmp_path, monkeypatch):
    account = _acct()
    first = tmp_path / "first.csv"
    second = tmp_path / "second.csv"
    first.write_text("value\n1\n")
    second.write_text("value\n2\n")
    original_inspect = datasets.inspect_dataset
    parsed_together = Barrier(2)

    def synchronized_inspect(*args, **kwargs):
        prepared = original_inspect(*args, **kwargs)
        parsed_together.wait(timeout=2)
        return prepared

    monkeypatch.setattr(datasets, "inspect_dataset", synchronized_inspect)

    def attempt(path):
        try:
            return datasets.register(account, "raced", str(path), "path", path.name)
        except HTTPException as exc:
            return exc.status_code

    with ThreadPoolExecutor(max_workers=2) as executor:
        outcomes = list(executor.map(attempt, (first, second)))

    assert sorted(409 if outcome == 409 else 200 for outcome in outcomes) == [200, 409]
    winner = datasets.query(account, "SELECT value FROM raced", ["raced"])["rows"]
    assert winner in ([[1]], [[2]])


def test_resync_keeps_old_version_queryable_until_atomic_switch(tmp_path):
    account = _acct()
    initial = tmp_path / "initial.csv"
    candidate = tmp_path / "candidate.csv"
    initial.write_text("value\n1\n")
    candidate.write_text("value\n2\n")
    datasets.register(account, "feed", str(initial), "url", "feed.csv", "https://example.test/feed", "csv")
    fetch_started = Event()
    allow_fetch = Event()
    cleaned: list[str] = []

    def fetcher(_url):
        fetch_started.set()
        assert allow_fetch.wait(timeout=2)
        return str(candidate)

    with ThreadPoolExecutor(max_workers=1) as executor:
        pending = executor.submit(
            datasets.resync,
            account,
            "feed",
            fetcher,
            None,
            "csv",
            cleaned.append,
        )
        assert fetch_started.wait(timeout=2)
        assert datasets.query(account, "SELECT value FROM feed", ["feed"])["rows"] == [[1]]
        allow_fetch.set()
        updated = pending.result(timeout=2)

    assert updated["location"] == str(candidate)
    assert updated["previous_location"] == str(initial)
    assert cleaned == []
    assert initial.exists()
    assert datasets.query(account, "SELECT value FROM feed", ["feed"])["rows"] == [[2]]


def test_url_reregistration_preserves_retired_version_until_durable_commit(tmp_path):
    account = _acct()
    initial = tmp_path / "initial.csv"
    replacement = tmp_path / "replacement.csv"
    initial.write_text("value\n1\n")
    replacement.write_text("value\n2\n")
    datasets.register(account, "feed", str(initial), "url", "feed.csv", "https://example.test/feed", "csv")

    updated = datasets.register(
        account,
        "feed",
        str(replacement),
        "url",
        "feed.csv",
        "https://example.test/feed",
        "csv",
    )

    assert updated["location"] == str(replacement)
    assert updated["previous_location"] == str(initial)
    assert initial.exists()
    assert datasets.query(account, "SELECT value FROM feed", ["feed"])["rows"] == [[2]]


def test_stale_location_deactivation_cannot_drop_same_name_replacement(tmp_path):
    account = _acct()
    old = tmp_path / "old.csv"
    current = tmp_path / "current.csv"
    old.write_text("value\n1\n")
    current.write_text("value\n2\n")
    datasets.register(account, "feed", str(old), "path", "old.csv")
    datasets.register(account, "feed", str(current), "path", "current.csv")

    assert datasets.deactivate_if_location(account, "feed", str(old)) is False
    assert datasets.query(account, "SELECT value FROM feed", ["feed"])["rows"] == [[2]]
    assert datasets.deactivate_if_location(account, "feed", str(current)) is True
    assert datasets.deactivate_if_location(account, "feed", str(current)) is False


def test_failed_durable_commit_can_roll_registry_back_to_preserved_old_version(tmp_path):
    account = _acct()
    initial = tmp_path / "durable.csv"
    candidate = tmp_path / "uncommitted.csv"
    initial.write_text("value\n1\n")
    candidate.write_text("value\n2\n")
    url = "https://example.test/feed"
    datasets.register(account, "feed", str(initial), "url", "feed.csv", url, "csv")

    switched = datasets.resync(
        account,
        "feed",
        fetcher=lambda _url: str(candidate),
        expected_format="csv",
        cleanup=lambda location: Path(location).unlink(missing_ok=True),
    )
    assert switched["previous_location"] == str(initial)
    assert initial.exists()

    rolled_back = datasets.register(account, "feed", str(initial), "url", "feed.csv", url, "csv")

    assert rolled_back["previous_location"] == str(candidate)
    assert datasets.query(account, "SELECT value FROM feed", ["feed"])["rows"] == [[1]]
    assert initial.exists()


def test_resync_cas_failure_discards_candidate_not_active_version(tmp_path):
    account = _acct()
    initial = tmp_path / "initial.csv"
    candidate = tmp_path / "candidate.csv"
    replacement = tmp_path / "replacement.csv"
    initial.write_text("value\n1\n")
    candidate.write_text("value\n2\n")
    replacement.write_text("value\n3\n")
    datasets.register(account, "feed", str(initial), "url", "feed.csv", "https://example.test/feed", "csv")
    fetch_started = Event()
    allow_fetch = Event()
    cleaned: list[str] = []

    def fetcher(_url):
        fetch_started.set()
        assert allow_fetch.wait(timeout=2)
        return str(candidate)

    with ThreadPoolExecutor(max_workers=1) as executor:
        pending = executor.submit(
            datasets.resync,
            account,
            "feed",
            fetcher,
            None,
            "csv",
            cleaned.append,
        )
        assert fetch_started.wait(timeout=2)
        datasets.register(account, "feed", str(replacement), "path", "replacement.csv")
        allow_fetch.set()
        with pytest.raises(HTTPException) as error:
            pending.result(timeout=2)

    assert error.value.status_code == 409
    assert cleaned == [str(candidate)]
    assert datasets.query(account, "SELECT value FROM feed", ["feed"])["rows"] == [[3]]


def test_extract_bounds_column_count_header_chars_and_reports_total_rows(tmp_path, monkeypatch):
    account = _acct()
    source = tmp_path / "wide.csv"
    source.write_text("abcdef,second,third\n1,2,3\n4,5,6\n")
    datasets.register(account, "wide", str(source), "path", "wide.csv")
    monkeypatch.setattr(datasets, "MAX_EXTRACT_COLUMNS", 2)
    monkeypatch.setattr(datasets, "MAX_EXTRACT_HEADER_CHARS", 3)

    result = datasets.extract(account, "wide", ["wide"], max_rows=1)

    assert result == {
        "columns": ["ab…"],
        "rows": [[1]],
        "row_count": 2,
        "total_row_count": 2,
        "returned_row_count": 1,
        "columns_truncated": True,
        "truncated": True,
    }
    assert sum(len(column) for column in result["columns"]) <= datasets.MAX_EXTRACT_HEADER_CHARS


def test_registration_bounds_column_count_and_column_names(tmp_path, monkeypatch):
    account = _acct()
    source = tmp_path / "wide.csv"
    source.write_text("first,second,third\n1,2,3\n")
    monkeypatch.setattr(datasets, "MAX_DATASET_COLUMNS", 2)

    with pytest.raises(HTTPException) as too_wide:
        datasets.register(account, "wide", str(source), "path", "wide.csv")

    assert too_wide.value.status_code == 413

    long_header = tmp_path / "long-header.csv"
    long_header.write_text(f"{'x' * 20}\nvalue\n")
    monkeypatch.setattr(datasets, "MAX_DATASET_COLUMNS", 500)
    monkeypatch.setattr(datasets, "MAX_DATASET_COLUMN_NAME_CHARS", 10)
    with pytest.raises(HTTPException) as long_name:
        datasets.register(_acct(), "long_header", str(long_header), "path", "long-header.csv")
    assert long_name.value.status_code == 413


def test_registry_preview_bounds_large_cells_before_response_materialization(tmp_path, monkeypatch):
    account = _acct()
    source = tmp_path / "large-preview.csv"
    source.write_text('text\n"' + ("x" * 10_000) + '"\n')
    monkeypatch.setattr(datasets, "MAX_PREVIEW_CELL_CHARS", 50)

    meta = datasets.register(account, "large_preview", str(source), "path", "large-preview.csv")

    assert meta["preview_truncated"] is True
    assert meta["preview"] == [["x" * 49 + "…"]]


def test_describe_bounds_top_values_and_total_text(tmp_path, monkeypatch):
    account = _acct()
    source = tmp_path / "large-describe.csv"
    source.write_text('text\n"' + ("z" * 5_000) + '"\n')
    monkeypatch.setattr(datasets, "MAX_DESCRIBE_VALUE_CHARS", 40)
    monkeypatch.setattr(datasets, "MAX_DESCRIBE_CHARS", 80)
    datasets.register(account, "large_describe", str(source), "path", "large-describe.csv")

    result = datasets.describe(account, "large_describe", ["large_describe"])

    value = result["columns"][0]["top_values"][0]["value"]
    assert len(value) <= 40
    assert value.endswith("…")
    assert result["truncated"] is True


def test_describe_casts_nested_values_to_a_bounded_string_inside_duckdb(tmp_path, monkeypatch):
    account = _acct()
    source = tmp_path / "nested.json"
    source.write_text('[{"payload":{"items":["' + ("n" * 5_000) + '"]}}]')
    monkeypatch.setattr(datasets, "MAX_DESCRIBE_VALUE_CHARS", 60)
    datasets.register(account, "nested", str(source), "url", "Nested", "https://example.test/nested", "json")

    result = datasets.describe(account, "nested", ["nested"])

    value = result["columns"][0]["top_values"][0]["value"]
    assert isinstance(value, str)
    assert len(value) == 60
    assert value.endswith("…")
    assert result["truncated"] is True


def test_catalog_is_exactly_scoped_and_aggregate_bounded(tmp_path, monkeypatch):
    account = _acct()
    _register_csv(tmp_path, account, "alpha", "very_long_alpha_column\n1\n")
    _register_csv(tmp_path, account, "beta", "very_long_beta_column\n2\n")
    _register_csv(tmp_path, account, "secret", "private\n3\n")
    one_item = datasets.catalog(account, ["alpha"])
    monkeypatch.setattr(datasets, "MAX_CATALOG_CHARS", len(str(one_item["datasets"][0])) + 20)

    result = datasets.catalog(account, ["beta", "alpha"])

    assert result["total"] == 2
    assert result["returned"] == 1
    assert result["omitted"] == 1
    assert result["truncated"] is True
    assert all(item["table"] != "secret" for item in result["datasets"])


def test_catalog_bounds_one_hundred_wide_registered_datasets(monkeypatch):
    account = _acct()
    columns = [{"name": f"column_{index}_" + "x" * 200, "type": "VARCHAR"} for index in range(40)]
    with datasets._account_lock(account):
        datasets._REGISTRY[account] = {
            f"table_{index}": {
                "name": f"table_{index}",
                "original_name": f"dataset-{index}.csv",
                "rows": 1,
                "columns": columns,
                "location": "/nonexistent",
            }
            for index in range(100)
        }
    monkeypatch.setattr(datasets, "MAX_CATALOG_CHARS", 50_000)
    try:
        result = datasets.catalog(account, [f"table_{index}" for index in range(100)])
    finally:
        with datasets._account_lock(account):
            datasets._REGISTRY.pop(account, None)

    encoded_chars = len(datasets.json.dumps(result["datasets"], ensure_ascii=False, separators=(",", ":")))
    assert encoded_chars <= datasets.MAX_CATALOG_CHARS + 2
    assert result["total"] == 100
    assert result["omitted"] > 0
    assert result["truncated"] is True


def test_real_xlsx_fixture_loads_offline_without_duckdb_extensions(tmp_path, monkeypatch):
    account = _acct()
    source = tmp_path / "fixture.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ledger"
    worksheet.append(["item", "amount"])
    worksheet.append(["rent", 1200.5])
    worksheet.append(["food", 42])
    workbook.save(source)
    workbook.close()
    original_loader = datasets.load_workbook
    original_reader = datasets._read_sql
    loaded_sources: list[str] = []

    def checked_loader(*args, **kwargs):
        assert kwargs["read_only"] is True
        assert kwargs["data_only"] is True
        assert kwargs["keep_links"] is False
        return original_loader(*args, **kwargs)

    def checked_reader(path, expected_format=None):
        loaded_sources.append(path)
        assert str(path).endswith(".csv")
        return original_reader(path, expected_format)

    monkeypatch.setattr(datasets, "load_workbook", checked_loader)
    monkeypatch.setattr(datasets, "_read_sql", checked_reader)

    meta = datasets.register(account, "ledger", str(source), "path", "fixture.xlsx")

    assert meta["rows"] == 2
    assert datasets.query(account, "SELECT item, amount FROM ledger ORDER BY item", ["ledger"])["rows"] == [
        ["food", 42.0],
        ["rent", 1200.5],
    ]
    assert len(loaded_sources) == 2


def test_xlsx_archive_limits_are_enforced_before_openpyxl(tmp_path, monkeypatch):
    source = tmp_path / "limited.xlsx"
    workbook = Workbook()
    workbook.active.append(["value"])
    workbook.active.append([1])
    workbook.save(source)
    workbook.close()
    monkeypatch.setattr(datasets, "MAX_XLSX_ARCHIVE_MEMBERS", 1)

    with pytest.raises(HTTPException) as error:
        datasets.register(_acct(), "limited", str(source), "path", "limited.xlsx")

    assert error.value.status_code == 413


def test_legacy_xls_is_explicitly_rejected(tmp_path):
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"not-an-xls-workbook")

    with pytest.raises(HTTPException) as error:
        datasets.register(_acct(), "legacy", str(source), "path", "legacy.xls")

    assert error.value.status_code == 422
    assert ".xls" in error.value.detail


def test_explicit_format_parses_extensionless_json_and_rejects_mismatch(tmp_path):
    source = tmp_path / "connector-cache"
    source.write_text('[{"value": 7}]')
    account = _acct()

    datasets.register(account, "events", str(source), "url", "Events", "https://example.test/events", "json")
    assert datasets.query(account, "SELECT value FROM events", ["events"])["rows"] == [[7]]

    with pytest.raises(HTTPException) as error:
        datasets.register(_acct(), "wrong", str(source), "url", "Wrong", "https://example.test/events", "csv")
    assert error.value.status_code == 422
